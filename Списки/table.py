import os
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Список файлов (дни)
files = ['1.08.xlsx', '2.08.xlsx', '3.08.xlsx', '4.08.xlsx']

# Листы (программы)
sheets = ['ПМ', 'ИВТ', 'ИТСС', 'ИБ']

# Все комбинации пересечений
combinations = [
    ('Только ПМ', {'ПМ'}), ('Только ИВТ', {'ИВТ'}), ('Только ИТСС', {'ИТСС'}), ('Только ИБ', {'ИБ'}),
    ('ПМ+ИВТ', {'ПМ', 'ИВТ'}), ('ПМ+ИТСС', {'ПМ', 'ИТСС'}), ('ПМ+ИБ', {'ПМ', 'ИБ'}),
    ('ИВТ+ИТСС', {'ИВТ', 'ИТСС'}), ('ИВТ+ИБ', {'ИВТ', 'ИБ'}), ('ИТСС+ИБ', {'ИТСС', 'ИБ'}),
    ('ПМ+ИВТ+ИТСС', {'ПМ', 'ИВТ', 'ИТСС'}), ('ПМ+ИВТ+ИБ', {'ПМ', 'ИВТ', 'ИБ'}),
    ('ПМ+ИТСС+ИБ', {'ПМ', 'ИТСС', 'ИБ'}), ('ИВТ+ИТСС+ИБ', {'ИВТ', 'ИТСС', 'ИБ'}),
    ('ПМ+ИВТ+ИТСС+ИБ', {'ПМ', 'ИВТ', 'ИТСС', 'ИБ'})
]

all_results = {}
unique_counts = {}   # Уникальные абитуриенты по дням
total_records = {}   # Общее количество записей (строк) по дням

# Анализ каждого файла
for file in files:
    if not os.path.exists(file):
        continue
    
    try:
        xlsx = pd.ExcelFile(file, engine='openpyxl')
        sheet_ids = {}
        day_total_records = 0
        
        # Загружаем данные с каждого листа
        for sheet in sheets:
            if sheet in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet, engine='openpyxl')
                day_total_records += len(df)  # Считаем ВСЕ строки на листе
                
                if 'ID' in df.columns:
                    ids = set(df['ID'].dropna().astype(str).str.strip().tolist())
                    sheet_ids[sheet] = ids
                else:
                    sheet_ids[sheet] = set()
            else:
                sheet_ids[sheet] = set()
        
        # Сохраняем общее количество записей за день
        total_records[file] = day_total_records
        
        # Собираем комбинации для каждого уникального ID
        id_to_sheets = defaultdict(set)
        for sheet_name, ids in sheet_ids.items():
            for id_val in ids:
                id_to_sheets[id_val].add(sheet_name)
        
        # Считаем точные совпадения комбинаций
        results = {}
        for combo_name, combo_sheets in combinations:
            count = sum(1 for sheets_set in id_to_sheets.values() if sheets_set == combo_sheets)
            results[combo_name] = count
        
        all_results[file] = results
        
        # Сохраняем количество уникальных абитуриентов
        unique_ids = set()
        for ids in sheet_ids.values():
            unique_ids.update(ids)
        unique_counts[file] = len(unique_ids)
        
    except Exception:
        continue

if not all_results:
    exit()

# Формируем транспонированную таблицу
df_original = pd.DataFrame([
    {'День': os.path.basename(f), **all_results[f]} for f in all_results
])

# Транспонируем: дни → столбцы, комбинации → строки
df_transposed = df_original.set_index('День').T
df_transposed.index.name = 'Комбинация программ'
df_transposed = df_transposed.reset_index()

# Переименовываем столбцы дней
day_columns = [col.replace('.xlsx', '') for col in df_transposed.columns if col != 'Комбинация программ']
rename_map = dict(zip(df_transposed.columns[1:], day_columns))
df_transposed = df_transposed.rename(columns=rename_map)

# Добавляем столбец ИТОГО по комбинациям (сумма по дням)
df_transposed['ИТОГО'] = df_transposed[day_columns].sum(axis=1)

# Строка 17: ИТОГО абитуриентов (уникальные)
итого_абитуриентов = {'Комбинация программ': 'ИТОГО абитуриентов'}
for day_file in files:
    day_name = day_file.replace('.xlsx', '')
    итого_абитуриентов[day_name] = unique_counts.get(day_file, 0)
итого_абитуриентов['ИТОГО'] = sum(итого_абитуриентов.get(d.replace('.xlsx', ''), 0) for d in files)

# Строка 18: ИТОГО записей (сумма строк со всех листов)
итого_записей = {'Комбинация программ': 'ИТОГО записей'}
for day_file in files:
    day_name = day_file.replace('.xlsx', '')
    итого_записей[day_name] = total_records.get(day_file, 0)
итого_записей['ИТОГО'] = sum(итого_записей.get(d.replace('.xlsx', ''), 0) for d in files)

# Формируем финальную таблицу
df_final = pd.concat([
    df_transposed,
    pd.DataFrame([итого_абитуриентов]),
    pd.DataFrame([итого_записей])
], ignore_index=True)

# Сохраняем в Excel
output_file = 'пересечения_итоги.xlsx'
df_final.to_excel(output_file, sheet_name='Сводная таблица', index=False)

# Форматирование
wb = load_workbook(output_file)
ws = wb['Сводная таблица']

# Автоширина столбцов
for col_idx, col in enumerate(df_final.columns, 1):
    max_len = max(
        df_final[col].astype(str).map(len).max(),
        len(str(col))
    ) + 2
    ws.column_dimensions[chr(64 + col_idx)].width = max_len

# Стили
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
center_align = Alignment(horizontal="center", vertical="center")

# Заголовок
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# Строка 17: ИТОГО абитуриентов — зелёный
строка_17 = ws.max_row - 1
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=строка_17, column=col_idx)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = center_align

# Строка 18: ИТОГО записей — оранжевый
строка_18 = ws.max_row
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=строка_18, column=col_idx)
    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    cell.font = Font(bold=True)
    cell.alignment = center_align

# Центрирование чисел
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.alignment = Alignment(horizontal="center")

wb.save(output_file)